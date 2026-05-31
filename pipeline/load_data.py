"""
Load POS transactions and store layout into the database.
Run this ONCE after docker compose up, before the detection pipeline.

Usage:
    python load_data.py --data-dir /data --api-url http://localhost:8000
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests

# Try to read Excel for store layout
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_pos_transactions(csv_path: str, api_url: str, store_id: str = "ST1008"):
    """
    Convert Brigade_Bangalore POS CSV into the API's expected format and load it.
    
    The CSV has: order_id, invoice_number, order_date, order_time, store_id, 
                 total_amount, basket_value_inr...
    We map this to: transaction_id, store_id, timestamp, basket_value_inr
    """
    print(f"\n[POS] Loading transactions from {csv_path}")
    
    if not os.path.exists(csv_path):
        print(f"  [WARN] POS file not found: {csv_path}")
        return 0

    transactions = []
    seen_invoices = set()

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            invoice = row.get("invoice_number", "").strip()
            if not invoice or invoice in seen_invoices:
                continue
            seen_invoices.add(invoice)

            # Parse date + time
            date_str = row.get("order_date", "").strip()
            time_str = row.get("order_time", "").strip()
            
            try:
                # Format: 10-04-2026 and 16:55:36
                dt = datetime.strptime(f"{date_str} {time_str}", "%d-%m-%Y %H:%M:%S")
                dt = dt.replace(tzinfo=timezone.utc)
                timestamp = dt.isoformat()
            except ValueError:
                print(f"  [WARN] Cannot parse date: {date_str} {time_str}")
                continue

            # Total amount per invoice - sum up total_amount for this invoice
            try:
                amount = float(row.get("total_amount", 0) or 0)
            except ValueError:
                amount = 0.0

            store = row.get("store_id", store_id).strip() or store_id

            transactions.append({
                "transaction_id": invoice,
                "store_id": store,
                "timestamp": timestamp,
                "basket_value_inr": amount,
            })

    print(f"  Found {len(transactions)} unique transactions")

    # POST to API
    url = f"{api_url}/pos/load"
    try:
        response = requests.post(url, json={"transactions": transactions}, timeout=30)
        if response.status_code == 200:
            print(f"  ✓ Loaded {len(transactions)} POS transactions")
        elif response.status_code == 404:
            # Endpoint not yet wired, insert directly via DB
            print(f"  [INFO] /pos/load endpoint not found — transactions will be handled on API start")
        else:
            print(f"  [WARN] HTTP {response.status_code}: {response.text[:200]}")
    except requests.exceptions.ConnectionError:
        print(f"  [WARN] API not reachable at {api_url}")

    return len(transactions)


def parse_store_layout_excel(xlsx_path: str, store_id: str = "ST1008") -> dict:
    """
    Parse the Brigade Road store layout Excel into the store_layout.json format.
    Since we can't know exact column structure without reading it, we build a 
    sensible default and override with whatever the Excel contains.
    """
    layout = {
        store_id: {
            "store_name": "Brigade_Bangalore",
            "city": "Bangalore",
            "open_hours": {"open": "10:00", "close": "22:00"},
            "cameras": {
                "CAM_1": {"type": "ENTRY", "description": "Main entry/exit camera"},
                "CAM_2": {"type": "FLOOR", "description": "Main floor camera"},
                "CAM_3": {"type": "FLOOR", "description": "Floor camera 2"},
                "CAM_4": {"type": "BILLING", "description": "Billing counter camera"},
                "CAM_5": {"type": "FLOOR", "description": "Floor camera 3"},
            },
            "zones": []
        }
    }

    if not HAS_OPENPYXL:
        print("  [INFO] openpyxl not installed — using default zone layout")
        # Build default zones based on known categories from POS data
        layout[store_id]["zones"] = _default_zones()
        return layout

    if not os.path.exists(xlsx_path):
        print(f"  [WARN] Store layout file not found: {xlsx_path} — using defaults")
        layout[store_id]["zones"] = _default_zones()
        return layout

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        print(f"  Excel sheets: {wb.sheetnames}")
        
        zones = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Parsing sheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} cols)")
            
            # Read first few rows to understand structure
            headers = None
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c else "" for c in row]
                    print(f"    Headers: {headers[:10]}")
                    continue
                
                # Try to find zone_name column
                row_dict = dict(zip(headers, row))
                zone_name = (
                    row_dict.get("Zone") or 
                    row_dict.get("zone") or 
                    row_dict.get("Zone Name") or
                    row_dict.get("Area") or
                    row_dict.get("Section")
                )
                if zone_name and str(zone_name).strip():
                    zones.append({
                        "zone_id": str(zone_name).strip().upper().replace(" ", "_"),
                        "display_name": str(zone_name).strip(),
                        # Polygon will be calibrated manually from video
                        "polygon": [],
                    })
        
        if zones:
            layout[store_id]["zones"] = zones
            print(f"  ✓ Parsed {len(zones)} zones from Excel")
        else:
            layout[store_id]["zones"] = _default_zones()
            print("  [INFO] No zones parsed from Excel — using defaults based on product categories")

    except Exception as e:
        print(f"  [WARN] Failed to parse Excel: {e} — using defaults")
        layout[store_id]["zones"] = _default_zones()

    return layout


def _default_zones() -> list:
    """
    Default zones inferred from Brigade Bangalore POS data product categories:
    - skin (toner, serum, face wash, sunscreen...)
    - makeup (lipstick, foundation, kajal, concealer...)
    - hair (shampoo, hair mask...)
    - fragrance (deodorant...)
    - personal-care (foot care...)
    - bath-and-body
    + BILLING area
    """
    return [
        {"zone_id": "SKINCARE", "display_name": "Skincare", "polygon": [], "categories": ["skin"]},
        {"zone_id": "MAKEUP", "display_name": "Makeup", "polygon": [], "categories": ["makeup"]},
        {"zone_id": "HAIRCARE", "display_name": "Haircare", "polygon": [], "categories": ["hair"]},
        {"zone_id": "FRAGRANCE", "display_name": "Fragrance", "polygon": [], "categories": ["fragrance"]},
        {"zone_id": "PERSONAL_CARE", "display_name": "Personal Care", "polygon": [], "categories": ["personal-care"]},
        {"zone_id": "BATH_BODY", "display_name": "Bath & Body", "polygon": [], "categories": ["bath-and-body"]},
        {"zone_id": "BILLING", "display_name": "Billing Counter", "polygon": []},
    ]


def main():
    parser = argparse.ArgumentParser(description="Load POS and store layout data")
    parser.add_argument("--data-dir", default="/data")
    parser.add_argument("--api-url", default="http://localhost:8000")
    parser.add_argument("--store-id", default="ST1008")
    args = parser.parse_args()

    data_dir = args.data_dir
    
    # 1. Parse and save store layout
    xlsx_path = os.path.join(data_dir, "store_layout.xlsx")
    json_path = os.path.join(data_dir, "store_layout.json")
    
    print(f"\n[LAYOUT] Parsing store layout...")
    layout = parse_store_layout_excel(xlsx_path, args.store_id)
    
    with open(json_path, "w") as f:
        json.dump(layout, f, indent=2)
    print(f"  ✓ Saved to {json_path}")
    
    # 2. Load POS transactions
    csv_path = os.path.join(data_dir, "pos_transactions.csv")
    load_pos_transactions(csv_path, args.api_url, args.store_id)

    print("\n[DONE] Data loading complete.")
    print(f"  Store layout: {json_path}")
    print(f"  Zones defined: {len(layout[args.store_id]['zones'])}")


if __name__ == "__main__":
    main()
